"""Parse a raw .eml file and extract its body text and attachments."""

import email
import email.header
import html
import io
import logging
import re
from dataclasses import dataclass

import openpyxl
import pypdf
from docx import Document as DocxDocument

logger = logging.getLogger(__name__)

TEXT_PREVIEW_CHARS = 500
BODY_MAX_CHARS = 2000
PDF_PREVIEW_PAGES = 2
PDF_PREVIEW_CHARS_PER_SIDE = 1000
DOCX_PREVIEW_CHARS_PER_SIDE = 1000
XLSX_PREVIEW_ROWS_PER_SHEET = 30
XLSX_PREVIEW_MAX_SHEETS = 3
XLSX_PREVIEW_MAX_CHARS = 2000

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class ParsedAttachment:
    filename: str
    mime_type: str
    size_bytes: int
    content: bytes
    text_preview: str | None


@dataclass
class ParsedEmail:
    """Parsed email containing the body text and all attachments."""
    body_text: str | None
    attachments: list[ParsedAttachment]


def _decode_filename(raw: str) -> str:
    parts = email.header.decode_header(raw)
    decoded = []
    for chunk, charset in parts:
        if isinstance(chunk, bytes):
            decoded.append(chunk.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(chunk)
    return "".join(decoded)


def _strip_html(html_text: str) -> str:
    """Crude HTML-to-text: strip tags, decode entities, collapse whitespace."""
    text = re.sub(r"<br\s*/?>", "\n", html_text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _extract_pdf_preview(content: bytes) -> str | None:
    """
    Extract a text preview from a PDF: first PDF_PREVIEW_PAGES pages and last
    PDF_PREVIEW_PAGES pages (skipping overlap on short PDFs). Each side is
    truncated to PDF_PREVIEW_CHARS_PER_SIDE chars. Returns None for scanned
    PDFs where no text can be extracted.
    """
    try:
        reader = pypdf.PdfReader(io.BytesIO(content))
    except Exception as e:
        logger.warning("Failed to open PDF: %s", e)
        return None

    n = len(reader.pages)
    if n == 0:
        return None

    def _read(indices: range) -> str:
        chunks: list[str] = []
        for i in indices:
            try:
                chunks.append(reader.pages[i].extract_text() or "")
            except Exception as e:
                logger.debug("Failed to extract page %d: %s", i, e)
        return "\n".join(c for c in chunks if c).strip()

    head_end = min(PDF_PREVIEW_PAGES, n)
    head = _read(range(0, head_end))[:PDF_PREVIEW_CHARS_PER_SIDE]

    if n <= PDF_PREVIEW_PAGES * 2:
        # Short PDF: head already covers (most of) it; skip tail to avoid duplication
        preview = head
    else:
        tail_start = n - PDF_PREVIEW_PAGES
        tail = _read(range(tail_start, n))[:PDF_PREVIEW_CHARS_PER_SIDE]
        preview = f"{head}\n[...]\n{tail}" if tail else head

    preview = preview.strip()
    return preview or None


def _extract_docx_preview(content: bytes) -> str | None:
    """
    Extract text from a DOCX: head and tail char windows of the concatenated
    paragraph text, separated by '[...]'. Returns None on parse failure or
    empty document.
    """
    try:
        doc = DocxDocument(io.BytesIO(content))
    except Exception as e:
        logger.warning("Failed to open DOCX: %s", e)
        return None

    paras = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    if not paras:
        return None

    full_text = "\n".join(paras).strip()
    if not full_text:
        return None

    if len(full_text) <= DOCX_PREVIEW_CHARS_PER_SIDE * 2:
        return full_text

    head = full_text[:DOCX_PREVIEW_CHARS_PER_SIDE]
    tail = full_text[-DOCX_PREVIEW_CHARS_PER_SIDE:]
    return f"{head}\n[...]\n{tail}"


def _extract_xlsx_preview(content: bytes) -> str | None:
    """
    Extract a tabular preview from an XLSX: for each of the first
    XLSX_PREVIEW_MAX_SHEETS sheets, render up to XLSX_PREVIEW_ROWS_PER_SHEET
    rows as tab-separated values. Total output capped at XLSX_PREVIEW_MAX_CHARS.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to open XLSX: %s", e)
        return None

    try:
        chunks: list[str] = []
        for sheet_name in wb.sheetnames[:XLSX_PREVIEW_MAX_SHEETS]:
            ws = wb[sheet_name]
            rows: list[str] = []
            truncated = False
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= XLSX_PREVIEW_ROWS_PER_SHEET:
                    truncated = True
                    break
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if not rows:
                continue
            block = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
            if truncated:
                block += "\n[...]"
            chunks.append(block)
    finally:
        wb.close()

    if not chunks:
        return None

    result = "\n\n".join(chunks)[:XLSX_PREVIEW_MAX_CHARS].strip()
    return result or None


def _extract_body(msg: email.message.Message) -> str | None:
    """
    Extract the plain-text body from an email message.
    Falls back to stripped HTML if no text/plain part is found.
    Truncates to BODY_MAX_CHARS.
    """
    plain_parts: list[str] = []
    html_parts: list[str] = []

    for part in msg.walk():
        # Skip attachments
        if part.get_content_disposition() in ("attachment", "inline") and part.get_filename():
            continue

        ct = part.get_content_type()
        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        charset = part.get_content_charset() or "utf-8"
        try:
            decoded = payload.decode(charset, errors="replace")
        except Exception:
            continue

        if ct == "text/plain":
            plain_parts.append(decoded)
        elif ct == "text/html":
            html_parts.append(decoded)

    if plain_parts:
        body = "\n".join(plain_parts)
    elif html_parts:
        body = _strip_html("\n".join(html_parts))
    else:
        return None

    body = body.strip()
    if not body:
        return None
    return body[:BODY_MAX_CHARS]


def parse_attachments(raw_bytes: bytes) -> list[ParsedAttachment]:
    """
    Extract all attachments from a raw RFC 822 email.
    Returns one ParsedAttachment per file found.
    """
    msg = email.message_from_bytes(raw_bytes)
    return _extract_attachments(msg)


def extract_text_preview(content: bytes, mime_type: str | None, filename: str) -> str | None:
    """
    Extract a text preview from a standalone file, mirroring the email-attachment
    pipeline. Used by the manual-upload flow so the same agent prompt works
    uniformly for both email and user-driven uploads.
    """
    if not content:
        return None
    lower_name = (filename or "").lower()
    mt = mime_type or ""

    if mt.startswith("text/"):
        try:
            return content.decode("utf-8", errors="replace")[:TEXT_PREVIEW_CHARS]
        except Exception:
            return None
    if mt == "application/pdf" or lower_name.endswith(".pdf"):
        return _extract_pdf_preview(content)
    if mt == DOCX_MIME or lower_name.endswith(".docx"):
        return _extract_docx_preview(content)
    if mt == XLSX_MIME or lower_name.endswith(".xlsx"):
        return _extract_xlsx_preview(content)
    return None


def parse_email(raw_bytes: bytes) -> ParsedEmail:
    """
    Parse a raw RFC 822 email and return both the body text and attachments.
    """
    msg = email.message_from_bytes(raw_bytes)
    body = _extract_body(msg)
    attachments = _extract_attachments(msg)
    logger.info("Parsed email: body=%d chars, %d attachment(s)",
                len(body) if body else 0, len(attachments))
    return ParsedEmail(body_text=body, attachments=attachments)


def _extract_attachments(msg: email.message.Message) -> list[ParsedAttachment]:
    """Extract all attachments from an already-parsed email.message.Message."""
    results: list[ParsedAttachment] = []

    for part in msg.walk():
        content_disposition = part.get_content_disposition()
        if content_disposition not in ("attachment", "inline"):
            continue

        raw_filename = part.get_filename()
        if not raw_filename:
            continue

        filename = _decode_filename(raw_filename)
        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        mime_type = part.get_content_type()

        text_preview = None
        lower_name = filename.lower()
        if mime_type.startswith("text/"):
            charset = part.get_content_charset() or "utf-8"
            try:
                text_preview = payload.decode(charset, errors="replace")[:TEXT_PREVIEW_CHARS]
            except Exception:
                pass
        elif mime_type == "application/pdf" or lower_name.endswith(".pdf"):
            text_preview = _extract_pdf_preview(payload)
        elif mime_type == DOCX_MIME or lower_name.endswith(".docx"):
            text_preview = _extract_docx_preview(payload)
        elif mime_type == XLSX_MIME or lower_name.endswith(".xlsx"):
            text_preview = _extract_xlsx_preview(payload)

        results.append(
            ParsedAttachment(
                filename=filename,
                mime_type=mime_type,
                size_bytes=len(payload),
                content=payload,
                text_preview=text_preview,
            )
        )
        logger.debug("Extracted attachment: %s (%s, %d bytes)", filename, mime_type, len(payload))

    logger.info("Parsed %d attachment(s) from .eml", len(results))
    return results
